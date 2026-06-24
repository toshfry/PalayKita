"""
PalayKita – Excel report generator
"""
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime


# ── Colour constants ─────────────────────────────────────────────────────────
C_GREEN       = '16A34A'
C_GREEN_LIGHT = 'DCFCE7'
C_GREEN_DARK  = '14532D'
C_AMBER       = 'D97706'
C_AMBER_LIGHT = 'FEF9C3'
C_RED         = 'DC2626'
C_RED_LIGHT   = 'FEE2E2'
C_GRAY_HEADER = 'F1F5F9'
C_GRAY_ROW    = 'F8FAFC'
C_WHITE       = 'FFFFFF'
C_TEXT        = '111827'
C_MUTED       = '6B7280'


def _D(v):
    return Decimal(str(v or 0))


def _peso(v):
    return f'₱{float(_D(v)):,.2f}'


def _thin():
    s = Side(style='thin', color='E5E7EB')
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _hdr_font(color=C_WHITE, size=10):
    return Font(name='Calibri', bold=True, size=size, color=color)


def _body_font(size=9, bold=False, color=C_TEXT):
    return Font(name='Calibri', size=size, bold=bold, color=color)


def _center(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)


def _right():
    return Alignment(horizontal='right', vertical='center')

def _payment_method_display(transaction):
    paid_total = _D(getattr(transaction, 'amount_paid', 0))
    if paid_total <= 0:
        return ''

    payment_records = list(getattr(transaction, 'payments', []) or [])
    payment_records.sort(key=lambda p: (
        getattr(p, 'payment_date', None) or datetime.min,
        getattr(p, 'created_at', None) or datetime.min,
        getattr(p, 'id', 0) or 0,
    ))

    recorded_total = sum(_D(getattr(p, 'amount', 0)) for p in payment_records)
    initial_total = paid_total - recorded_total
    methods = []

    if initial_total > 0 and getattr(transaction, 'payment_method', None):
        methods.append(str(transaction.payment_method).strip())

    for payment in payment_records:
        method = str(getattr(payment, 'payment_method', '') or '').strip()
        if method:
            methods.append(method)

    if not methods and getattr(transaction, 'payment_method', None):
        methods.append(str(transaction.payment_method).strip())

    unique = []
    for method in methods:
        if method and method not in unique:
            unique.append(method)
    return ' / '.join(unique)


def _payment_notes_display(transaction, symbol='₱'):
    notes = []
    if getattr(transaction, 'notes', None):
        note = str(transaction.notes).strip()
        if note:
            notes.append(note)

    payment_records = list(getattr(transaction, 'payments', []) or [])
    payment_records.sort(key=lambda p: (
        getattr(p, 'payment_date', None) or datetime.min,
        getattr(p, 'created_at', None) or datetime.min,
        getattr(p, 'id', 0) or 0,
    ))

    for payment in payment_records:
        note = str(getattr(payment, 'notes', '') or '').strip()
        if not note:
            continue
        pdate = getattr(payment, 'payment_date', None)
        if pdate and hasattr(pdate, 'strftime'):
            pdate = pdate.strftime('%Y-%m-%d')
        method = str(getattr(payment, 'payment_method', 'Payment') or 'Payment').strip()
        amount = f"{symbol}{float(_D(getattr(payment, 'amount', 0))):,.2f}"
        details = ' - '.join(part for part in [pdate, method, amount] if part)
        notes.append(f"Payment Note ({details}): {note}" if details else f"Payment Note: {note}")

    return ' | '.join(notes)



# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel_report(filepath, transactions, settings, report_type,
                           start_date, end_date, business_name):
    wb = Workbook()
    ws = wb.active
    ws.title = 'PalayKita Report'

    _write_title(ws, business_name, report_type, start_date, end_date)
    summary = _compute_summary(transactions)
    _write_summary(ws, summary, len(transactions))
    _write_table(ws, transactions, summary, getattr(settings, 'currency_symbol', '₱') or '₱')
    _write_footer(ws, settings, transactions)
    _set_col_widths(ws)

    wb.save(filepath)


# ── Sections ──────────────────────────────────────────────────────────────────

NCOLS = 17   # number of data columns

def _col_range(ws, row, col_start, col_end, **kw):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row, c)
        for attr, val in kw.items():
            setattr(cell, attr, val)


def _write_title(ws, biz, rtype, s_date, e_date):
    # Row 1 – Business Name
    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    c = ws['A1']
    c.value     = biz
    c.font      = Font(name='Calibri', bold=True, size=20, color=C_GREEN_DARK)
    c.alignment = _center()
    c.fill      = _fill(C_GREEN_LIGHT)
    ws.row_dimensions[1].height = 38

    # Row 2 – Report type
    labels = {'daily':'Daily Report','weekly':'Weekly Report',
              'monthly':'Monthly Report','custom':'Custom Date Range Report'}
    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    c = ws['A2']
    c.value     = f'🌾  {labels.get(rtype, "Report")}'
    c.font      = Font(name='Calibri', bold=True, size=13, color=C_GREEN)
    c.alignment = _center()
    c.fill      = _fill('F0FDF4')
    ws.row_dimensions[2].height = 26

    # Row 3 – Date range
    if s_date == e_date:
        date_label = s_date.strftime('%B %d, %Y')
    else:
        date_label = f'{s_date.strftime("%B %d, %Y")} – {e_date.strftime("%B %d, %Y")}'
    ws.merge_cells(f'A3:{get_column_letter(NCOLS)}3')
    c = ws['A3']
    c.value     = date_label
    c.font      = Font(name='Calibri', size=11, color=C_MUTED, italic=True)
    c.alignment = _center()
    c.fill      = _fill('F0FDF4')
    ws.row_dimensions[3].height = 20

    # Row 4 – spacer
    ws.row_dimensions[4].height = 8


def _compute_summary(txns):
    return {
        'count'           : len(txns),
        'kilos'           : sum(_D(t.kilos_milled)    for t in txns),
        'gross_fees'      : sum(_D(t.gross_fee)        for t in txns),
        'chaff_deductions': sum(_D(t.chaff_deduction)  for t in txns),
        'net_income'      : sum(_D(t.net_amount)       for t in txns),
        'cash_collected'  : sum(_D(t.amount_paid)      for t in txns),
        'unpaid_balance'  : sum(_D(t.balance)          for t in txns),
    }


def _write_summary(ws, s, count):
    START = 5
    # Header
    ws.merge_cells(f'A{START}:{get_column_letter(NCOLS)}{START}')
    c = ws.cell(START, 1, '  SUMMARY')
    c.font      = _hdr_font(size=11)
    c.fill      = _fill(C_GREEN)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[START].height = 24

    rows = [
        ('Total Transactions',      str(s['count'])),
        ('Total Kilos Milled',      f'{float(s["kilos"]):,.2f} kg'),
        ('Gross Milling Fees',      _peso(s['gross_fees'])),
        ('Rice Chaff Deductions',   _peso(s['chaff_deductions'])),
        ('Net Income',              _peso(s['net_income'])),
        ('Cash Collected',          _peso(s['cash_collected'])),
        ('Outstanding Balance',     _peso(s['unpaid_balance'])),
    ]

    for i, (label, value) in enumerate(rows):
        r = START + 1 + i
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f'A{r}:D{r}')
        lc = ws.cell(r, 1, f'  {label}')
        lc.font      = _body_font(bold=True, size=10)
        lc.fill      = _fill(C_GRAY_HEADER)
        lc.alignment = Alignment(horizontal='left', vertical='center')

        vc = ws.cell(r, 5, value)
        vc.font      = _body_font(bold=True, size=10,
                                  color=C_GREEN_DARK if label == 'Net Income' else C_TEXT)
        vc.fill      = _fill(C_GRAY_HEADER)
        vc.alignment = _right()

        for col in range(1, NCOLS + 1):
            ws.cell(r, col).border = _thin()
            if col not in (1, 2, 3, 4, 5):
                ws.cell(r, col).fill = _fill(C_GRAY_HEADER)

    # Spacer row
    sp = START + len(rows) + 1
    ws.row_dimensions[sp].height = 10


HEADER_ROW = 14

def _write_table(ws, txns, summary, symbol='₱'):
    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        'Txn No.', 'Date', 'Time', 'Customer / Owner', 'Contact No.',
        'Kilos', 'Rate/kg', 'Gross Fee',
        'Chaff kg', 'Chaff Rate', 'Chaff Deduction',
        'Net Amount', 'Amount Paid', 'Balance',
        'Status', 'Method', 'Notes'
    ]

    ws.row_dimensions[HEADER_ROW].height = 24
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(HEADER_ROW, col, hdr)
        c.font      = _hdr_font()
        c.fill      = _fill(C_GREEN)
        c.alignment = _center(wrap=True)
        c.border    = _thin()

    ws.freeze_panes = f'A{HEADER_ROW + 1}'

    STATUS_COLORS = {
        'Paid'   : (C_GREEN_LIGHT, C_GREEN_DARK),
        'Unpaid' : (C_RED_LIGHT,   C_RED),
        'Partial': (C_AMBER_LIGHT, C_AMBER),
    }

    CURRENCY_COLS = {8, 11, 12, 13, 14}

    for ri, t in enumerate(txns, 1):
        r   = HEADER_ROW + ri
        bg  = C_WHITE if ri % 2 == 0 else C_GRAY_ROW
        ws.row_dimensions[r].height = 18

        customer = t.customer_name or 'Walk-in / Not specified'
        time_str = t.created_at.strftime('%I:%M %p') if t.created_at else ''

        row_vals = [
            t.transaction_number,
            t.transaction_date.strftime('%Y-%m-%d') if t.transaction_date else '',
            time_str,
            customer,
            t.contact_number or '',
            float(_D(t.kilos_milled)),
            float(_D(t.milling_rate_per_kg)),
            float(_D(t.gross_fee)),
            float(_D(t.chaff_kilos))        if t.has_chaff_deduction else 0,
            float(_D(t.chaff_rate_per_kg))  if t.has_chaff_deduction else 0,
            float(_D(t.chaff_deduction)),
            float(_D(t.net_amount)),
            float(_D(t.amount_paid)),
            float(_D(t.balance)),
            t.payment_status,
            _payment_method_display(t),
            _payment_notes_display(t, symbol),
        ]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(r, ci, val)
            c.border = _thin()

            if ci == 15:   # Status column
                sc = STATUS_COLORS.get(val, (C_WHITE, C_TEXT))
                c.fill      = _fill(sc[0])
                c.font      = _body_font(bold=True, size=9, color=sc[1])
                c.alignment = _center()
            else:
                c.fill = _fill(bg)
                c.font = _body_font()
                if ci in CURRENCY_COLS:
                    c.number_format = '#,##0.00'
                    c.alignment     = _right()
                elif ci == 6:
                    c.number_format = '#,##0.00'
                    c.alignment     = _right()
                else:
                    c.alignment = Alignment(vertical='center')

    # ── Totals row ────────────────────────────────────────────────────────────
    if txns:
        tr = HEADER_ROW + len(txns) + 1
        ws.row_dimensions[tr].height = 22
        ws.merge_cells(f'A{tr}:E{tr}')
        ws.cell(tr, 1, '  TOTALS').font      = _hdr_font(size=10)
        ws.cell(tr, 1).fill                   = _fill(C_GREEN)
        ws.cell(tr, 1).alignment              = Alignment(horizontal='left', vertical='center')

        total_map = {
            6 : float(summary['kilos']),
            8 : float(summary['gross_fees']),
            11: float(summary['chaff_deductions']),
            12: float(summary['net_income']),
            13: float(summary['cash_collected']),
            14: float(summary['unpaid_balance']),
        }
        for col in range(1, NCOLS + 1):
            c = ws.cell(tr, col)
            c.fill   = _fill(C_GREEN)
            c.border = _thin()
            if col in total_map:
                c.value         = total_map[col]
                c.font          = _hdr_font(size=10)
                c.number_format = '#,##0.00'
                c.alignment     = _right()


def _write_footer(ws, settings, txns):
    if not (settings and settings.receipt_footer):
        return
    footer_row = HEADER_ROW + len(txns) + 3
    ws.merge_cells(f'A{footer_row}:{get_column_letter(NCOLS)}{footer_row}')
    c = ws.cell(footer_row, 1, settings.receipt_footer)
    c.font      = Font(name='Calibri', italic=True, size=9, color=C_MUTED)
    c.alignment = _center()


def _set_col_widths(ws):
    widths = [20, 12, 10, 26, 16, 9, 9, 13, 9, 11, 15, 13, 13, 12, 10, 12, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
