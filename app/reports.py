from datetime import date, timedelta, datetime
from pathlib import Path
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app

from app.models import MillingTransaction, CommercialTransaction
from app.utils import get_settings
from app.calculations import money


# ── Report colors matched to the uploaded sample ──────────────────────────────
C_GREEN = "16A34A"
C_GREEN_LIGHT = "DCFCE7"
C_GREEN_DARK = "14532D"
C_AMBER = "D97706"
C_AMBER_LIGHT = "FEF9C3"
C_RED = "DC2626"
C_RED_LIGHT = "FEE2E2"
C_GRAY_HEADER = "F1F5F9"
C_GRAY_ROW = "F8FAFC"
C_WHITE = "FFFFFF"
C_TEXT = "111827"
C_MUTED = "6B7280"

NCOLS = 17
HEADER_ROW = 14


def _D(value):
    return Decimal(str(value or 0))


def _peso(value, symbol="₱"):
    return f"{symbol}{float(_D(value)):,.2f}"


def _thin():
    side = Side(style="thin", color="E5E7EB")
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _hdr_font(color=C_WHITE, size=10):
    return Font(name="Calibri", bold=True, size=size, color=color)


def _body_font(size=9, bold=False, color=C_TEXT, italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _right():
    return Alignment(horizontal="right", vertical="center")


def _peso_format(symbol):
    return f'"{symbol}"#,##0.00'


def _query_transactions(start_date, end_date, payment_status="all"):
    query = MillingTransaction.query.filter(
        MillingTransaction.transaction_date >= start_date,
        MillingTransaction.transaction_date <= end_date,
    )

    if payment_status in ["Paid", "Partial", "Unpaid"]:
        query = query.filter(MillingTransaction.payment_status == payment_status)

    return query.order_by(
        MillingTransaction.transaction_date.asc(),
        MillingTransaction.created_at.asc(),
    ).all()


def _query_commercial_transactions(start_date, end_date, customer_id=None, payment_status="all"):
    query = CommercialTransaction.query.filter(
        CommercialTransaction.transaction_date >= start_date,
        CommercialTransaction.transaction_date <= end_date,
    )

    if customer_id:
        query = query.filter(CommercialTransaction.customer_id == customer_id)

    if payment_status in ["Paid", "Partial", "Unpaid"]:
        query = query.filter(CommercialTransaction.payment_status == payment_status)

    return query.order_by(
        CommercialTransaction.transaction_date.asc(),
        CommercialTransaction.created_at.asc(),
    ).all()


def _summary(transactions):
    return {
        "total_transactions": len(transactions),
        "total_kilos": sum(money(t.kilos_milled) for t in transactions),
        "gross": sum(money(t.gross_fee) for t in transactions),
        "chaff": sum(money(t.chaff_deduction) for t in transactions),
        "net": sum(money(t.net_amount) for t in transactions),
        "cash": sum(money(t.amount_paid) for t in transactions),
        "balance": sum(money(t.balance) for t in transactions),
    }


def _commercial_summary(transactions):
    return {
        "total_transactions": len(transactions),
        "total_sacks": sum(money(t.number_of_sacks) for t in transactions),
        "total_amount": sum(money(t.total_amount) for t in transactions),
        "total_paid": sum(money(t.amount_paid) for t in transactions),
        "total_balance": sum(money(t.balance) for t in transactions),
    }


def _payment_method_display(transaction):
    """
    Show the actual method(s) used, including additional payments recorded later.
    """
    paid_total = money(transaction.amount_paid)
    if paid_total <= 0:
        return ""

    payment_records = list(getattr(transaction, "payments", []) or [])
    payment_records.sort(key=lambda p: (
        p.payment_date or date.min,
        p.created_at or datetime.min,
        p.id or 0,
    ))

    recorded_payment_total = sum(money(p.amount) for p in payment_records)
    initial_payment_total = paid_total - recorded_payment_total

    methods = []
    if initial_payment_total > 0 and transaction.payment_method:
        methods.append(str(transaction.payment_method).strip())

    for payment in payment_records:
        if payment.payment_method:
            methods.append(str(payment.payment_method).strip())

    if not methods and transaction.payment_method:
        methods.append(str(transaction.payment_method).strip())

    unique_methods = []
    for method in methods:
        if method and method not in unique_methods:
            unique_methods.append(method)

    return " / ".join(unique_methods)


def _payment_notes_display(transaction, symbol="₱"):
    """
    Include transaction notes and notes entered when paying unpaid balances.
    """
    notes = []

    if transaction.notes:
        original_note = str(transaction.notes).strip()
        if original_note:
            notes.append(original_note)

    payment_records = list(getattr(transaction, "payments", []) or [])
    payment_records.sort(key=lambda p: (
        p.payment_date or date.min,
        p.created_at or datetime.min,
        p.id or 0,
    ))

    for payment in payment_records:
        payment_note = str(payment.notes or "").strip()
        if not payment_note:
            continue

        payment_date = payment.payment_date.strftime("%Y-%m-%d") if payment.payment_date else ""
        payment_method = str(payment.payment_method or "Payment").strip()
        payment_amount = f"{symbol}{money(payment.amount):,.2f}"
        details = " - ".join(part for part in [payment_date, payment_method, payment_amount] if part)
        notes.append(f"Payment Note ({details}): {payment_note}" if details else f"Payment Note: {payment_note}")

    return " | ".join(notes)


def _write_title(ws, business_name, report_type, start_date, end_date):
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    cell = ws["A1"]
    cell.value = business_name
    cell.font = Font(name="Calibri", bold=True, size=20, color=C_GREEN_DARK)
    cell.alignment = _center()
    cell.fill = _fill(C_GREEN_LIGHT)
    ws.row_dimensions[1].height = 38

    labels = {
        "daily": "Daily Report",
        "weekly": "Weekly Report",
        "monthly": "Monthly Report",
        "custom": "Custom Date Range Report",
    }
    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    cell = ws["A2"]
    cell.value = f"🌾  {labels.get(report_type, 'Report')}"
    cell.font = Font(name="Calibri", bold=True, size=13, color=C_GREEN)
    cell.alignment = _center()
    cell.fill = _fill("F0FDF4")
    ws.row_dimensions[2].height = 26

    if start_date == end_date:
        date_label = start_date.strftime("%B %d, %Y")
    else:
        date_label = f"{start_date.strftime('%B %d, %Y')} – {end_date.strftime('%B %d, %Y')}"

    ws.merge_cells(f"A3:{get_column_letter(NCOLS)}3")
    cell = ws["A3"]
    cell.value = date_label
    cell.font = Font(name="Calibri", size=11, color=C_MUTED, italic=True)
    cell.alignment = _center()
    cell.fill = _fill("F0FDF4")
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 8


def _write_summary(ws, summary, symbol):
    start = 5
    ws.merge_cells(f"A{start}:{get_column_letter(NCOLS)}{start}")
    cell = ws.cell(start, 1, "SUMMARY")
    cell.font = _hdr_font(size=11)
    cell.fill = _fill(C_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start].height = 24

    rows = [
        ("Total Transactions", str(summary["total_transactions"])),
        ("Total Kilos Milled", f'{float(summary["total_kilos"]):,.2f} kg'),
        ("Gross Milling Fees", _peso(summary["gross"], symbol)),
        ("Rice Chaff Deductions", _peso(summary["chaff"], symbol)),
        ("Net Income", _peso(summary["net"], symbol)),
        ("Cash Collected", _peso(summary["cash"], symbol)),
        ("Outstanding Balance", _peso(summary["balance"], symbol)),
    ]

    for index, (label, value) in enumerate(rows, start=start + 1):
        ws.row_dimensions[index].height = 20
        ws.merge_cells(f"A{index}:D{index}")

        label_cell = ws.cell(index, 1, label)
        label_cell.font = _body_font(bold=True, size=10)
        label_cell.fill = _fill(C_GRAY_HEADER)
        label_cell.alignment = Alignment(horizontal="left", vertical="center")

        value_cell = ws.cell(index, 5, value)
        value_cell.font = _body_font(
            bold=True,
            size=10,
            color=C_GREEN_DARK if label == "Net Income" else C_TEXT,
        )
        value_cell.fill = _fill(C_GRAY_HEADER)
        value_cell.alignment = _right()

        for col in range(1, NCOLS + 1):
            c = ws.cell(index, col)
            c.border = _thin()
            if col not in (1, 2, 3, 4, 5):
                c.fill = _fill(C_GRAY_HEADER)

    ws.row_dimensions[start + len(rows) + 1].height = 10


def _write_table(ws, transactions, summary, symbol):
    headers = [
        "Txn No.", "Date", "Time", "Customer / Owner", "Contact No.",
        "Kilos", "Rate/kg", "Gross Fee", "Chaff kg", "Chaff Rate",
        "Chaff Deduction", "Net Amount", "Amount Paid", "Balance",
        "Status", "Method", "Notes",
    ]

    ws.row_dimensions[HEADER_ROW].height = 24
    for col, header in enumerate(headers, 1):
        cell = ws.cell(HEADER_ROW, col, header)
        cell.font = _hdr_font()
        cell.fill = _fill(C_GREEN)
        cell.alignment = _center(wrap=True)
        cell.border = _thin()

    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    status_colors = {
        "Paid": (C_GREEN_LIGHT, C_GREEN_DARK),
        "Unpaid": (C_RED_LIGHT, C_RED),
        "Partial": (C_AMBER_LIGHT, C_AMBER),
    }
    currency_cols = {8, 11, 12, 13, 14}
    numeric_cols = {6, 7, 9, 10}

    for idx, transaction in enumerate(transactions, 1):
        row = HEADER_ROW + idx
        bg = C_WHITE if idx % 2 == 0 else C_GRAY_ROW
        ws.row_dimensions[row].height = 18

        created_at = transaction.created_at
        time_str = created_at.strftime("%I:%M %p") if created_at else ""

        values = [
            transaction.transaction_number,
            transaction.transaction_date.strftime("%Y-%m-%d") if transaction.transaction_date else "",
            time_str,
            transaction.customer_name or "Walk-in / Not specified",
            transaction.contact_number or "",
            float(money(transaction.kilos_milled)),
            float(money(transaction.milling_rate_per_kg)),
            float(money(transaction.gross_fee)),
            float(money(transaction.chaff_kilos)) if transaction.has_chaff_deduction else 0,
            float(money(transaction.chaff_rate_per_kg)) if transaction.has_chaff_deduction else 0,
            float(money(transaction.chaff_deduction)),
            float(money(transaction.net_amount)),
            float(money(transaction.amount_paid)),
            float(money(transaction.balance)),
            transaction.payment_status,
            _payment_method_display(transaction),
            _payment_notes_display(transaction, symbol),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = _thin()

            if col == 15:
                fill_color, font_color = status_colors.get(value, (C_WHITE, C_TEXT))
                cell.fill = _fill(fill_color)
                cell.font = _body_font(bold=True, size=9, color=font_color)
                cell.alignment = _center()
            else:
                cell.fill = _fill(bg)
                cell.font = _body_font()
                if col in currency_cols:
                    cell.number_format = _peso_format(symbol)
                    cell.alignment = _right()
                elif col in numeric_cols:
                    cell.number_format = "#,##0.00"
                    cell.alignment = _right()
                elif col == 17:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="center")

    if transactions:
        total_row = HEADER_ROW + len(transactions) + 1
        ws.row_dimensions[total_row].height = 22
        ws.merge_cells(f"A{total_row}:E{total_row}")
        total_label = ws.cell(total_row, 1, "TOTALS")
        total_label.font = _hdr_font(size=10)
        total_label.fill = _fill(C_GREEN)
        total_label.alignment = Alignment(horizontal="left", vertical="center")

        totals = {
            6: float(summary["total_kilos"]),
            8: float(summary["gross"]),
            11: float(summary["chaff"]),
            12: float(summary["net"]),
            13: float(summary["cash"]),
            14: float(summary["balance"]),
        }
        for col in range(1, NCOLS + 1):
            cell = ws.cell(total_row, col)
            cell.fill = _fill(C_GREEN)
            cell.border = _thin()
            if col in totals:
                cell.value = totals[col]
                cell.font = _hdr_font(size=10)
                cell.number_format = _peso_format(symbol) if col != 6 else "#,##0.00"
                cell.alignment = _right()


def _write_footer(ws, settings, transactions):
    footer = getattr(settings, "receipt_footer", "") or ""
    if not footer:
        return

    footer_row = HEADER_ROW + len(transactions) + 3
    ws.merge_cells(f"A{footer_row}:{get_column_letter(NCOLS)}{footer_row}")
    cell = ws.cell(footer_row, 1, footer)
    cell.font = _body_font(size=9, color=C_MUTED, italic=True)
    cell.alignment = _center()


def _set_column_widths(ws):
    widths = [20, 12, 10, 26, 16, 9, 9, 13, 9, 11, 15, 13, 13, 12, 10, 12, 32]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _style_page(ws):
    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4


def _write_commercial_sheet(ws, transactions, summary, settings):
    symbol = settings.currency_symbol or "â‚±"
    title = settings.commercial_receipt_label or "Commercial Transactions"
    headers = [
        "Txn No.", "Date", "Time", "Customer", "Contact No.", "Sacks",
        "Price/Sack", "Total Amount", "Amount Paid", "Balance", "Status", "Notes",
    ]

    ws.merge_cells("A1:L1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name="Calibri", bold=True, size=18, color=C_GREEN_DARK)
    cell.alignment = _center()
    cell.fill = _fill(C_GREEN_LIGHT)
    ws.row_dimensions[1].height = 34

    summary_rows = [
        ("Total Transactions", summary["total_transactions"]),
        ("Total Sacks", float(summary["total_sacks"])),
        ("Total Amount", float(summary["total_amount"])),
        ("Total Paid", float(summary["total_paid"])),
        ("Total Balance", float(summary["total_balance"])),
    ]

    for row, (label, value) in enumerate(summary_rows, start=3):
        label_cell = ws.cell(row, 1, label)
        value_cell = ws.cell(row, 2, value)
        label_cell.font = _body_font(bold=True)
        value_cell.font = _body_font(bold=True)
        if row == 4:
            value_cell.number_format = "#,##0.00"
        elif row >= 5:
            value_cell.number_format = _peso_format(symbol)

    header_row = 10
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = _hdr_font()
        cell.fill = _fill(C_GREEN)
        cell.alignment = _center(wrap=True)
        cell.border = _thin()

    status_colors = {
        "Paid": (C_GREEN_LIGHT, C_GREEN_DARK),
        "Unpaid": (C_RED_LIGHT, C_RED),
        "Partial": (C_AMBER_LIGHT, C_AMBER),
    }
    currency_cols = {7, 8, 9, 10}

    for idx, transaction in enumerate(transactions, 1):
        row = header_row + idx
        bg = C_WHITE if idx % 2 == 0 else C_GRAY_ROW
        created_at = transaction.created_at
        time_str = created_at.strftime("%I:%M %p") if created_at else ""
        customer = transaction.customer

        values = [
            transaction.transaction_number,
            transaction.transaction_date.strftime("%Y-%m-%d") if transaction.transaction_date else "",
            time_str,
            customer.name if customer else "",
            customer.contact_number if customer else "",
            float(money(transaction.number_of_sacks)),
            float(money(transaction.price_per_sack)),
            float(money(transaction.total_amount)),
            float(money(transaction.amount_paid)),
            float(money(transaction.balance)),
            transaction.payment_status,
            transaction.notes or "",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row, col, value)
            cell.border = _thin()
            if col == 11:
                fill_color, font_color = status_colors.get(value, (C_WHITE, C_TEXT))
                cell.fill = _fill(fill_color)
                cell.font = _body_font(bold=True, size=9, color=font_color)
                cell.alignment = _center()
            else:
                cell.fill = _fill(bg)
                cell.font = _body_font()
                if col in currency_cols:
                    cell.number_format = _peso_format(symbol)
                    cell.alignment = _right()
                elif col == 6:
                    cell.number_format = "#,##0.00"
                    cell.alignment = _right()
                elif col == 12:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="center")

    if transactions:
        total_row = header_row + len(transactions) + 1
        ws.merge_cells(f"A{total_row}:E{total_row}")
        total_label = ws.cell(total_row, 1, "TOTALS")
        total_label.font = _hdr_font(size=10)
        total_label.fill = _fill(C_GREEN)

        totals = {
            6: float(summary["total_sacks"]),
            8: float(summary["total_amount"]),
            9: float(summary["total_paid"]),
            10: float(summary["total_balance"]),
        }
        for col in range(1, len(headers) + 1):
            cell = ws.cell(total_row, col)
            cell.fill = _fill(C_GREEN)
            cell.border = _thin()
            if col in totals:
                cell.value = totals[col]
                cell.font = _hdr_font(size=10)
                cell.number_format = "#,##0.00" if col == 6 else _peso_format(symbol)
                cell.alignment = _right()

    widths = [20, 12, 10, 26, 16, 10, 12, 14, 14, 14, 10, 34]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    _style_page(ws)


def generate_excel_report(report_type, start_date, end_date, payment_status="all"):
    settings = get_settings()
    symbol = settings.currency_symbol or "₱"
    transactions = _query_transactions(start_date, end_date, payment_status)
    summary = _summary(transactions)
    commercial_transactions = _query_commercial_transactions(start_date, end_date, payment_status=payment_status)
    commercial_summary = _commercial_summary(commercial_transactions)

    wb = Workbook()
    ws = wb.active
    ws.title = "PalayKita Report"

    _write_title(ws, settings.business_name, report_type, start_date, end_date)
    _write_summary(ws, summary, symbol)
    _write_table(ws, transactions, summary, symbol)
    _write_footer(ws, settings, transactions)
    _set_column_widths(ws)
    _style_page(ws)

    commercial_ws = wb.create_sheet("Commercial Sales")
    _write_commercial_sheet(commercial_ws, commercial_transactions, commercial_summary, settings)

    generated_time = datetime.now().strftime("%H-%M-%S")
    filename = f"PalayKita_{report_type.title()}_Report_{start_date.strftime('%Y-%m-%d')}"
    if start_date != end_date:
        filename += f"_to_{end_date.strftime('%Y-%m-%d')}"
    filename += f"_{generated_time}.xlsx"

    if report_type == "daily":
        folder = current_app.config["DAILY_REPORT_DIR"]
    elif report_type == "weekly":
        folder = current_app.config["WEEKLY_REPORT_DIR"]
    elif report_type == "monthly":
        folder = current_app.config["MONTHLY_REPORT_DIR"]
    else:
        folder = current_app.config["CUSTOM_REPORT_DIR"]

    Path(folder).mkdir(parents=True, exist_ok=True)
    output_path = Path(folder) / filename
    wb.save(output_path)
    return output_path


def generate_commercial_customer_report(start_date, end_date, customer_id=None, payment_status="all"):
    settings = get_settings()
    transactions = _query_commercial_transactions(start_date, end_date, customer_id, payment_status)
    summary = _commercial_summary(transactions)

    wb = Workbook()
    ws = wb.active
    ws.title = "Commercial Report"
    _write_commercial_sheet(ws, transactions, summary, settings)

    generated_time = datetime.now().strftime("%H-%M-%S")
    filename = f"PalayKita_Commercial_Report_{start_date.strftime('%Y-%m-%d')}"
    if start_date != end_date:
        filename += f"_to_{end_date.strftime('%Y-%m-%d')}"
    filename += f"_{generated_time}.xlsx"

    folder = current_app.config["COMMERCIAL_REPORT_DIR"]
    Path(folder).mkdir(parents=True, exist_ok=True)
    output_path = Path(folder) / filename
    wb.save(output_path)
    return output_path


def daily_range(day=None):
    selected = day or date.today()
    return selected, selected


def weekly_range(day=None):
    selected = day or date.today()
    start = selected - timedelta(days=selected.weekday())
    end = start + timedelta(days=6)
    return start, end


def monthly_range(day=None):
    selected = day or date.today()
    start = selected.replace(day=1)
    if selected.month == 12:
        end = selected.replace(year=selected.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end = selected.replace(month=selected.month + 1, day=1) - timedelta(days=1)
    return start, end
